# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from django.shortcuts import render
from django.template import RequestContext
from django.http import HttpResponse
from django.template.loader import get_template
from django.shortcuts import redirect

from learn import models
from .forms import LoginForm
#from learn import forms
from .models import Work

from django import forms


from django.http import HttpResponseRedirect
from django.urls import reverse

from .forms import UserInfo
from openpyxl import Workbook
from openpyxl import  load_workbook
from django.core.mail import EmailMultiAlternatives
# Create your views here.

i=0
j=0
def user_list(request):
	global i
	f_obj = UserInfo()
	if request.method ==  'POST':
		#wb = Workbook()
		wb = load_workbook('roll number.xlsx')
		ws = wb.active
		ws.title = "兼职报名统计"
		number = ws.cell(row=2,column=1)
		number.value = "序号"
		#name = ws.cell('A2')
		name = ws.cell(row=2,column=2)
		name.value = "名字"
		#phone = ws.cell('B2')
		phone = ws.cell(row=2,column=3)
		phone.value = "手机号"
		sex = ws.cell(row=2, column=4)
		#sex = ws.cell('C2')
		sex.value = "性别"
		#school = ws.cell('D2')
		school = ws.cell(row=2, column=5)
		school.value = "学校"


		user_input_obj = UserInfo(request.POST)
		if user_input_obj.is_valid():
			i +=1
			data = user_input_obj.clean()
			f_name = user_input_obj.cleaned_data['name']
			f_phone = user_input_obj.cleaned_data['phone']
			f_sex = user_input_obj.cleaned_data['sex']
			f_school = user_input_obj.cleaned_data['school']
			ws.append({'A':i,'B':f_name,'C':f_phone,'D':f_sex,'E':f_school})
			wb.save('roll number.xlsx')
			print f_name
		else:
			error_msg = user_input_obj.errors
			return render(request,'user_list.html',{'obj':user_input_obj,'errors':error_msg})
	return render(request,'user_list.html',context={'obj':f_obj})



def index(request):
	if 'username' in request.session:
		username = request.session['username']
		useremail = request.session['useremail']
	template = get_template('index.html')
	html = template.render(locals())
	return HttpResponse(html)



def contact(request):
	template = get_template('contact.html')
	html = template.render(locals())
	return HttpResponse(html)

def xls_process(UserInfo):
	global j
	j +=1
	user_input_obj = UserInfo
	wb = load_workbook('roll number.xlsx')
	ws = wb.active
	ws.title = "兼职报名统计"
	number = ws.cell(row=2, column=1)
	number.value = "序号"
	# name = ws.cell('A2')
	name = ws.cell(row=2, column=2)
	name.value = "名字"
	# phone = ws.cell('B2')
	phone = ws.cell(row=2, column=3)
	phone.value = "手机号"
	sex = ws.cell(row=2, column=4)
	# sex = ws.cell('C2')
	sex.value = "性别"
	# school = ws.cell('D2')
	school = ws.cell(row=2, column=5)
	school.value = "学校"
	roll_content = ws.cell(row=2, column=6)
	roll_content.value = "报名信息"
	f_name = user_input_obj.cleaned_data['name']
	f_phone = user_input_obj.cleaned_data['phone']
	f_sex = user_input_obj.cleaned_data['sex']
	f_school = user_input_obj.cleaned_data['school']
	f_roll = user_input_obj.cleaned_data['content']
	ws.append({'A': j, 'B': f_name, 'C': f_phone, 'D': f_sex, 'E': f_school,'F':f_roll})
	wb.save('roll number.xlsx')
	subject, form_email, to = 'hello', 'wangyu_0898@sina.com', '270179788@qq.com'
	text_content = 'this is a test to send msg from django'
	html_content = u'<b>激活链接：</b><a href="http://www.baidu.com">http:www.baidu.com</a>'
	msg = EmailMultiAlternatives(subject, text_content, form_email, [to])
	msg.attach_alternative(html_content, 'text/html')
	msg.attach_file('roll number.xlsx')
	msg.send()

def notic(request):

	#words = 'World!'
	works = Work.objects.all()
	f_obj = UserInfo()
	if request.method ==  'POST':
		user_obj = UserInfo(request.POST)
		if user_obj.is_valid():
			data = user_obj.clean()
			xls_process(user_obj)
			print data
		else:
			erro_msg = user_obj.errors
			return render(request,'user_list.html',{'obj':user_obj,'errors':erro_msg})
	#return render(request,'user_list.html',context={'obj':f_obj})
	return render(request, 'notic.html', context={'notices': works,'obj':f_obj})
#def student(request):
#	students = Student.objects.all()
#	return render(request,'student.html',context={'students':students})

#def student(request):
#    students = Student.objects.all()
#    if request.method == 'POST':
#        form = StudentForm(request.POST)
#        if form.is_valid():
#            cleaned_data = form.cleaned_data
#            student = Student()
#            student.name = cleaned_data['name']
#            student.sex = cleaned_data['sex']
#            student.email = cleaned_data['email']
#            student.profession = cleaned_data['profession']
#            student.qq = cleaned_data['qq']
#            student.phone = cleaned_data['phone']
#            student.save()
#            return HttpResponseRedirect(reverse('student'))
#    else:
##        form = StudentForm()
#
#    context = {
#        'students': students,
#        'form': form,
#    }
#    return render(request, 'student.html', context=context)
def login(request):
	obj_login = LoginForm()
	#message = '请您登录'
	if request.method == 'POST':
		login_form = LoginForm(request.POST)
		print('enter post')
		if login_form.is_valid():
			login_name = login_form.cleaned_data['username']
			login_password = login_form.cleaned_data['password']
			#login_name=request.POST['username'].strip()
			#login_password=request.POST['password']
			print('the user login name is %s',login_name)
			print('the user login password is %s', login_password)
			#user = authenticate(username=login_name, password=login_password)
			try:
				#temp_user = User.objects.get(name= wangyu)
				#userAll = models.User.objects.all()

				#user = User.objects.get(name='wangyu') 不正确语法，所以错误
				user = models.User.objects.get(name= login_name )
				print('find the temp user  %s', user)
				if user.password == login_password:
					response = redirect('/')
					request.session['username'] = user.name
					request.session['useremail'] = user.email
					print('yes it is member')
					return redirect('/')
				else:
					message = "密码错误，请再检查一次"
			except:
				print('exception')
				message ="目前无法登陆"
		else:
			message = "请检查输入的字段内容"
	else:
		login_form = LoginForm()

	template = get_template('login.html')
	#request_context = RequestContext(request)
	#request_context.push(locals())
	#html = template.render(locals())
	return render(request, 'login.html', context={'obj': obj_login,'message' : message})
	#return HttpResponse(html)